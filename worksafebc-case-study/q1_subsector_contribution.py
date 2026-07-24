"""
WorkSafeBC Case Study - Question 1
Subsector contribution to the BC provincial injury rate, and the effect of a
targeted 10% injury-rate reduction in the top-3 contributing subsectors.

Data source: WorkSafeBC public Power BI dashboard "Provincial Overview
(past 10 years)", Injury Rate tab, extracted via right-click > "Show as a table"
for each of the 24 ratable subsectors (sectors 70-76). Data as of 2026-06-30.

Analysis window: 2019-2023 (last five COMPLETED injury years relative to the exam
framing, which references the 2.08 rate 'as of 2023' and a 'below 2.0 by 2024'
goal). DISCLAIMER: the live dashboard now also carries 2024-2025 data; a re-run on
2021-2025 is provided separately for comparison.

Author: <candidate name>   |   AI assistance: see AI_ATTRIBUTION.md
Run:  python q1_subsector_contribution.py  ->  q1_outputs.xlsx + q1_subsector_data.csv
"""
import csv

YEARS = [2016,2017,2018,2019,2020,2021,2022,2023,2024,2025]
WINDOW = [2019,2020,2021,2022,2023]
WIN_IDX = [YEARS.index(y) for y in WINDOW]

SUB = {
 "7010 Agriculture":                 dict(tlc=[541,553,620,572,712,810,639,593,473,568],   py=[21943,22273,22662,24896,24412,26409,25418,23807,23789,25790]),
 "7020 Fishing":                     dict(tlc=[195,130,160,156,140,132,134,106,86,65],      py=[None]*10),
 "7030 Forestry":                    dict(tlc=[767,747,786,654,533,548,660,505,467,416],    py=[15976,16681,16996,15282,13468,14536,13875,13120,12368,11414]),
 "7040 Oil, Gas or Mineral Resources":dict(tlc=[192,238,260,250,250,336,371,324,308,387],   py=[17766,20638,19939,20746,18234,20319,25491,25814,24079,24065]),
 "7110 Food and Beverage Products":  dict(tlc=[1236,1293,1335,1392,1352,1473,1367,1326,1215,1194], py=[32062,33209,33520,33946,32282,33163,33293,32103,33100,32686]),
 "7120 Metal and Non-Metallic Mineral":dict(tlc=[1883,1893,2052,1946,1598,1820,1829,1744,1693,1650], py=[48404,51026,53186,54083,48876,51280,52169,52741,55807,53800]),
 "7130 Petroleum/Coal/Rubber/Chemical":dict(tlc=[474,457,447,498,392,431,435,402,372,341],  py=[23736,25032,25709,27026,26407,26905,29068,31272,28181,26511]),
 "7140 Wood and Paper Products":     dict(tlc=[1660,1595,1522,1453,1268,1406,1362,1178,1053,990], py=[54202,54855,54083,51720,44460,48734,48160,43213,41413,39464]),
 "7150 Other Products (nes)":        dict(tlc=[727,630,696,666,550,625,573,522,496,457],    py=[46451,46249,46461,45713,41809,45574,50054,49523,51507,53640]),
 "7210 General Construction":        dict(tlc=[6632,7024,7388,7046,6353,6905,6638,6911,6722,6355], py=[159704,178629,181755,191842,185729,198718,215773,224977,216419,213005]),
 "7220 Heavy Construction":          dict(tlc=[86,89,94,96,86,94,171,192,114,102],          py=[2901,2865,3120,4239,5016,6792,7496,7556,7346,7309]),
 "7230 Road Construction/Maintenance":dict(tlc=[311,380,407,385,364,421,337,368,340,334],   py=[12110,14151,14880,14866,14661,14959,15488,14937,14109,14328]),
 "7310 Warehousing":                 dict(tlc=[244,265,225,192,201,213,172,169,177,270],    py=[3683,4016,3935,4178,3791,3842,4250,4200,4268,5283]),
 "7320 Transportation & Related Svcs":dict(tlc=[4104,4326,4402,4604,4177,4891,4905,4688,4874,5265], py=[93724,98070,101409,106152,100242,104670,112263,115284,116919,127261]),
 "7410 Retail":                      dict(tlc=[5121,5009,5179,5332,4901,5420,5139,4942,4696,4969], py=[241047,247478,245969,247546,229304,234264,240120,240952,242909,237964]),
 "7420 Wholesale":                   dict(tlc=[1082,1095,1168,1180,1023,1176,1117,1130,1046,1060], py=[68255,69105,68727,70328,65571,69379,70213,66083,64710,61934]),
 "7530 Public Administration":       dict(tlc=[2970,2989,3199,3256,3062,3479,3730,3396,3310,3520], py=[123023,124275,126683,129300,121190,127765,129099,129834,134608,131980]),
 "7610 Tourism and Hospitality":     dict(tlc=[6125,6292,6544,6637,4331,5073,5687,5991,5901,6141], py=[393720,404618,410971,418066,321900,342486,379794,386765,391420,388897]),
 "7620 Business Services":           dict(tlc=[1611,1616,1655,1725,1373,1575,1534,1505,1407,1452], py=[346507,353488,359350,364510,343708,378610,388167,389603,401709,405929]),
 "7630 Professional/Scientific/Tech":dict(tlc=[2031,2176,2205,2301,1925,2315,2164,1929,2008,1908], py=[228232,238404,249431,262064,263106,308427,324712,319568,316627,313435]),
 "7640 Other Services (nes)":        dict(tlc=[5205,5381,5589,5676,5009,5548,5587,5399,5484,5443], py=[214294,221538,229122,235027,216805,230855,241295,235160,238404,246055]),
 "7650 Education":                   dict(tlc=[2097,2293,2387,2597,2024,2727,4552,3128,3053,3230], py=[142373,147034,151481,157121,156474,150042,146612,161594,167591,165566]),
 "7660 Health Care & Social Services":dict(tlc=[8869,9037,8979,10049,11260,11940,16620,12698,12849,12983], py=[248484,249650,256249,265692,269865,280332,291174,313675,329876,339488]),
 "7670 Utilities":                   dict(tlc=[338,295,260,248,213,210,229,206,210,162],    py=[34870,33821,32364,32005,29233,28079,27104,28122,26482,26562]),
}
PROV = dict(tlc=[49091,50328,51719,53011,47982,53688,60367,53702,53124,53962],
            py=[2232194,2311580,2364366,2424707,2248686,2399246,2520024,2579485,2620092,2642700])

def wsum(arr):
    return sum(v for v in (arr[i] for i in WIN_IDX) if v is not None)

def analyse():
    prov_rate_2023 = PROV["tlc"][YEARS.index(2023)] / PROV["py"][YEARS.index(2023)] * 100.0
    prov_rate_pooled = wsum(PROV["tlc"]) / wsum(PROV["py"]) * 100.0
    total_sub_claims = sum(wsum(d["tlc"]) for d in SUB.values())
    rows = []
    for name, d in SUB.items():
        c = wsum(d["tlc"]); p = wsum(d["py"]) if d["py"][0] is not None else None
        share = c / total_sub_claims
        own_rate = (c / p * 100.0) if p else None
        contrib = share * prov_rate_2023
        rows.append(dict(name=name, claims=c, py=p, own_rate=own_rate, share=share, contrib=contrib))
    rows.sort(key=lambda r: r["contrib"], reverse=True)
    top3 = rows[:3]
    top3_share = sum(r["share"] for r in top3)
    new_rate = prov_rate_2023 * (1 - top3_share * 0.10)
    x_needed = (1 - 2.0/prov_rate_2023) / top3_share
    return dict(prov_rate_2023=prov_rate_2023, prov_rate_pooled=prov_rate_pooled, rows=rows,
                top3=top3, top3_share=top3_share, new_rate=new_rate, x_needed=x_needed)

def write_csv(res, path="q1_subsector_data.csv"):
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["rank","subsector","claims_2019_2023","person_years_2019_2023",
                    "own_injury_rate","claim_share_pct","contribution_to_prov_rate"])
        for i, r in enumerate(res["rows"], 1):
            w.writerow([i, r["name"], r["claims"], r["py"] or "",
                        f"{r['own_rate']:.2f}" if r["own_rate"] else "",
                        f"{r['share']*100:.2f}", f"{r['contrib']:.4f}"])

def write_xlsx(res, path="q1_outputs.xlsx"):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        print("openpyxl not installed; skipping .xlsx (CSV still written).")
        return
    wb = Workbook(); ws = wb.active; ws.title = "Q1 Contribution"
    ws.append(["Rank","Subsector","TLC 2019-2023","PY 2019-2023","Own injury rate",
               "Claim share %","Contribution (anchored 2.08)"])
    for c in ws[1]: c.font = Font(bold=True)
    for i, r in enumerate(res["rows"], 1):
        ws.append([i, r["name"], r["claims"], r["py"] or "n/a",
                   round(r["own_rate"],2) if r["own_rate"] else "n/a",
                   round(r["share"]*100,2), round(r["contrib"],4)])
    ws2 = wb.create_sheet("Intervention"); ws2.append(["Metric","Value"])
    for c in ws2[1]: c.font = Font(bold=True)
    ws2.append(["Provincial injury rate 2023 (official)", round(res["prov_rate_2023"],3)])
    ws2.append(["Provincial injury rate 2019-2023 pooled", round(res["prov_rate_pooled"],3)])
    ws2.append(["Top 3 contributors", ", ".join(r["name"] for r in res["top3"])])
    ws2.append(["Top 3 combined claim share %", round(res["top3_share"]*100,1)])
    ws2.append(["New provincial rate after 10% cut to top-3", round(res["new_rate"],3)])
    ws2.append(["Below 2.0 target met?", "YES (thin margin)" if res["new_rate"]<2.0 else "NO"])
    ws2.append(["Reduction on top-3 to reach exactly 2.0", f"{res['x_needed']*100:.1f}%"])
    wb.save(path)

if __name__ == "__main__":
    res = analyse()
    print(f"Province injury rate 2023 (official): {res['prov_rate_2023']:.3f}")
    print("Top 3 by contribution: " + ", ".join(f"{r['name']} ({r['contrib']:.4f})" for r in res["top3"]))
    print(f"Top 3 combined claim share: {res['top3_share']*100:.1f}%")
    print(f"New provincial rate after 10% cut to top-3: {res['new_rate']:.3f}")
    print(f"Below-2.0 target met: {res['new_rate']<2.0}")
    print(f"Reduction on top-3 to reach exactly 2.0: {res['x_needed']*100:.1f}%")
    write_csv(res); write_xlsx(res)
    print("Wrote q1_subsector_data.csv and q1_outputs.xlsx")
